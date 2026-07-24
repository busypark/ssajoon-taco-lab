"""
algorithm_solution_score_matrix.batch1.xlsx vs .batch2.xlsx 비교.

batch1: solution 1개씩 호출 (비교 문맥 없음)
batch2: 같은 문제 내 solution 2개씩 묶어서 호출 (일부 상호 비교 가능)

같은 problem_solution_key 행을 조인해서:
  1. primary_algorithm 일치율 + 안 맞을 때 어떤 태그로 바뀌었는지
  2. confidence 분포 변화 (batch1->batch2 크로스탭)
  3. 태그별 점수 평균 절대 차이 (어느 태그가 batch에 따라 가장 크게 흔들리는지)
  4. compared_with_other_solutions 가 "비교 불가" 상투문구인 비율

사용법: python compare_batches.py
출력: stdout 요약 + output/batch1_vs_batch2_diff.csv (행 단위 상세 비교)
"""
import csv
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

import schema

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"
BATCH1_XLSX = OUTPUT_DIR / "algorithm_solution_score_matrix.batch1.xlsx"
BATCH2_XLSX = OUTPUT_DIR / "algorithm_solution_score_matrix.batch2.xlsx"
DIFF_CSV = OUTPUT_DIR / "batch1_vs_batch2_diff.csv"

UNIQUE_TAGS = schema.DEDUPED_ALGO_TAGS  # 120개, 중복 제거된 태그만 (점수 비교용)
NO_COMPARISON_MARKER = "비교 불가"


def load_all_score(xlsx_path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["ALL_ALGORITHM_SCORE"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    out = {}
    for row in rows:
        d = dict(zip(header, row))
        out[d["problem_solution_key"]] = d
    return out


def load_recheck(xlsx_path: Path) -> dict[tuple[str, str], dict]:
    """key: (problem_solution_key, source_broad_tag) — 한 solution이 여러 broad_tag로 recheck될 수 있음."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["BROAD_TAG_RECHECK_SCORE"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    out = {}
    for row in rows:
        d = dict(zip(header, row))
        out[(d["problem_solution_key"], d["source_broad_tag"])] = d
    return out


def to_int(v) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def main():
    b1 = load_all_score(BATCH1_XLSX)
    b2 = load_all_score(BATCH2_XLSX)

    keys_b1, keys_b2 = set(b1), set(b2)
    joined_keys = sorted(keys_b1 & keys_b2)

    print(f"batch1 rows: {len(b1)}, batch2 rows: {len(b2)}, joined (both present): {len(joined_keys)}")
    print(f"only in batch1: {sorted(keys_b1 - keys_b2)}")
    print(f"only in batch2: {sorted(keys_b2 - keys_b1)}")
    print()

    # 1. primary_algorithm 일치율
    match_count = 0
    primary_changes = Counter()
    for key in joined_keys:
        p1, p2 = b1[key]["primary_algorithm"], b2[key]["primary_algorithm"]
        if p1 == p2:
            match_count += 1
        else:
            primary_changes[(p1, p2)] += 1

    print(f"=== 1. primary_algorithm 일치율: {match_count}/{len(joined_keys)} "
          f"({match_count / len(joined_keys) * 100:.1f}%) ===")
    print("가장 흔한 변경 (batch1 -> batch2):")
    for (p1, p2), n in primary_changes.most_common(15):
        print(f"  {p1} -> {p2}: {n}건")
    print()

    # 2. confidence 크로스탭
    conf_crosstab = Counter()
    for key in joined_keys:
        conf_crosstab[(b1[key]["confidence"], b2[key]["confidence"])] += 1
    conf_levels = ["HIGH", "MEDIUM", "LOW"]

    print("=== 2. confidence 분포 변화 (batch1 x batch2) ===")
    header_line = "batch1\\batch2".ljust(10) + "".join(c.ljust(10) for c in conf_levels)
    print(header_line)
    for c1 in conf_levels:
        row_str = c1.ljust(10) + "".join(str(conf_crosstab[(c1, c2)]).ljust(10) for c2 in conf_levels)
        print(row_str)
    b1_conf_totals = Counter(b1[k]["confidence"] for k in joined_keys)
    b2_conf_totals = Counter(b2[k]["confidence"] for k in joined_keys)
    print(f"batch1 합계: {dict(b1_conf_totals)}")
    print(f"batch2 합계: {dict(b2_conf_totals)}")
    print()

    # 3. 태그별 평균 절대 점수 차이
    tag_abs_diffs = defaultdict(list)
    all_diffs = []
    for key in joined_keys:
        for tag in UNIQUE_TAGS:
            d = abs(to_int(b1[key].get(tag)) - to_int(b2[key].get(tag)))
            tag_abs_diffs[tag].append(d)
            all_diffs.append(d)

    print(f"=== 3. 태그별 점수 평균 절대 차이 (전체 {len(UNIQUE_TAGS)}개 태그 평균: "
          f"{sum(all_diffs) / len(all_diffs):.2f}점) ===")
    tag_avg = sorted(((tag, sum(vs) / len(vs)) for tag, vs in tag_abs_diffs.items()),
                      key=lambda x: -x[1])
    print("가장 많이 흔들리는 태그 top 15:")
    for tag, avg in tag_avg[:15]:
        print(f"  {tag}: 평균 {avg:.1f}점 차이")
    print()

    # 4. compared_with_other_solutions 상투문구 비율
    b1_boilerplate = sum(1 for k in joined_keys if NO_COMPARISON_MARKER in (b1[k].get("compared_with_other_solutions") or ""))
    b2_boilerplate = sum(1 for k in joined_keys if NO_COMPARISON_MARKER in (b2[k].get("compared_with_other_solutions") or ""))
    print(f"=== 4. compared_with_other_solutions 가 '{NO_COMPARISON_MARKER}' 상투문구인 비율 ===")
    print(f"batch1: {b1_boilerplate}/{len(joined_keys)} ({b1_boilerplate / len(joined_keys) * 100:.1f}%)")
    print(f"batch2: {b2_boilerplate}/{len(joined_keys)} ({b2_boilerplate / len(joined_keys) * 100:.1f}%)")
    print()

    # 행 단위 상세 diff CSV
    with open(DIFF_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "problem_solution_key", "primary_algorithm_b1", "primary_algorithm_b2", "primary_match",
            "confidence_b1", "confidence_b2", "max_tag_diff", "max_tag_diff_tag",
            "avg_tag_diff", "compared_b1_is_boilerplate", "compared_b2_is_boilerplate",
        ])
        for key in joined_keys:
            diffs = {tag: abs(to_int(b1[key].get(tag)) - to_int(b2[key].get(tag))) for tag in UNIQUE_TAGS}
            max_tag = max(diffs, key=diffs.get)
            w.writerow([
                key, b1[key]["primary_algorithm"], b2[key]["primary_algorithm"],
                b1[key]["primary_algorithm"] == b2[key]["primary_algorithm"],
                b1[key]["confidence"], b2[key]["confidence"],
                diffs[max_tag], max_tag, round(sum(diffs.values()) / len(diffs), 2),
                NO_COMPARISON_MARKER in (b1[key].get("compared_with_other_solutions") or ""),
                NO_COMPARISON_MARKER in (b2[key].get("compared_with_other_solutions") or ""),
            ])
    print(f"행 단위 상세 비교는 {DIFF_CSV} 에 저장됨.")
    print()

    q3_stats = compare_q3set(b1, b2, joined_keys)
    print()
    compare_recheck()
    print()
    print_summary_table(match_count, q3_stats, len(joined_keys))


def q3_selected_tags(row: dict) -> set[str]:
    """0점을 제외한 나머지 점수들의 Q3(75퍼센타일) 이상인 태그만 모은 집합.
    태그가 1개뿐이면 그 자체가 Q3(=그 값)이므로 그 태그만 포함된다."""
    scores = {tag: to_int(row.get(tag)) for tag in UNIQUE_TAGS}
    nonzero = [v for v in scores.values() if v > 0]
    if not nonzero:
        return set()
    if len(nonzero) == 1:
        q3 = nonzero[0]
    else:
        q3 = statistics.quantiles(nonzero, n=4, method="inclusive")[2]
    return {tag for tag, v in scores.items() if v > 0 and v >= q3}


def q3_set_score(set1: set[str], set2: set[str]) -> float:
    """작은 집합을 A, 큰 집합을 B로 두고 intersection/|A| - 0.1*|B-A|. 둘 다 비어있으면 1점, A가 비어있으면 0점."""
    if not set1 and not set2:
        return 1.0
    A, B = (set1, set2) if len(set1) <= len(set2) else (set2, set1)
    if not A:
        return 0.0
    base = len(A & B) / len(A)
    penalty = 0.1 * len(B - A)
    return max(0.0, base - penalty)


def compare_q3set(b1: dict, b2: dict, joined_keys: list[str]):
    print("=== 6. Q3 기반 적응형 집합 비교 (문제마다 집합 크기가 다름) ===")
    print("점수 = |교집합|/|작은집합| - 0.1*|큰집합에만 있는 개수| (0~1, 완전포함이면 페널티만큼 감점)")

    scores = []
    size1_list, size2_list = [], []
    low_examples = []

    for key in joined_keys:
        s1 = q3_selected_tags(b1[key])
        s2 = q3_selected_tags(b2[key])
        size1_list.append(len(s1))
        size2_list.append(len(s2))
        s = q3_set_score(s1, s2)
        scores.append(s)
        if s < 1.0:
            low_examples.append((key, s, s1, s2))

    avg = sum(scores) / len(scores)
    full_match = sum(1 for s in scores if s >= 0.999)
    zero_match = sum(1 for s in scores if s <= 0.001)
    partial_match = len(scores) - full_match - zero_match

    print(f"평균 점수: {avg:.3f} / 1.0  (완전일치 1.0점: {full_match}건, 부분일치: {partial_match}건, "
          f"불일치 0점: {zero_match}건 / 총 {len(joined_keys)}건)")
    print(f"batch1 집합 평균 크기: {sum(size1_list)/len(size1_list):.2f}개, "
          f"batch2 집합 평균 크기: {sum(size2_list)/len(size2_list):.2f}개")
    print(f"참고: primary_algorithm 단순 일치율은 "
          f"{sum(1 for k in joined_keys if b1[k]['primary_algorithm'] == b2[k]['primary_algorithm']) / len(joined_keys) * 100:.1f}% 였음")

    low_examples.sort(key=lambda x: x[1])
    print(f"\n점수가 가장 낮은 예시 (최대 7개, 1.0 미만 총 {len(low_examples)}건):")
    for key, s, s1, s2 in low_examples[:7]:
        print(f"  {key}: {s:.2f}점 | batch1={sorted(s1)} / batch2={sorted(s2)}")

    return {"avg": avg, "full_match": full_match, "partial_match": partial_match, "zero_match": zero_match}


SUMMARY_CSV = OUTPUT_DIR / "batch1_vs_batch2_summary.csv"


def print_summary_table(primary_match_count: int, q3_stats: dict, n: int):
    primary_full = primary_match_count
    primary_mismatch = n - primary_match_count
    primary_avg = primary_full / n

    q3_full, q3_partial, q3_zero = q3_stats["full_match"], q3_stats["partial_match"], q3_stats["zero_match"]
    q3_avg = q3_stats["avg"]

    rows = [
        ("완전 일치 (건수)", str(primary_full), str(q3_full)),
        ("완전 일치 (비율)", f"{primary_full / n * 100:.1f}%", f"{q3_full / n * 100:.1f}%"),
        ("부분 일치 (건수)", "N/A (이진 지표라 없음)", str(q3_partial)),
        ("부분 일치 (비율)", "N/A", f"{q3_partial / n * 100:.1f}%"),
        ("불일치 (건수)", str(primary_mismatch), str(q3_zero)),
        ("불일치 (비율)", f"{primary_mismatch / n * 100:.1f}%", f"{q3_zero / n * 100:.1f}%"),
        ("평균 점수 (0~1)", f"{primary_avg:.3f}", f"{q3_avg:.3f}"),
    ]

    print(f"=== 7. 요약 비교표: primary만 비교 vs Q3 기반 적응형 집합 비교 (총 {n}건 기준) ===")
    col_w = [20, 26, 30]
    header = ["지표", "primary만 비교", "Q3 적응형 집합 비교"]
    print("".join(h.ljust(w) for h, w in zip(header, col_w)))
    for r in rows:
        print("".join(str(c).ljust(w) for c, w in zip(r, col_w)))

    with open(SUMMARY_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"\n요약 테이블은 {SUMMARY_CSV} 에도 저장됨.")


def compare_recheck():
    r1 = load_recheck(BATCH1_XLSX)
    r2 = load_recheck(BATCH2_XLSX)

    keys_r1, keys_r2 = set(r1), set(r2)
    joined = sorted(keys_r1 & keys_r2)

    print("=== 5. BROAD_TAG_RECHECK_SCORE 비교 (problem_solution_key + source_broad_tag 기준) ===")
    print(f"batch1 recheck 행: {len(r1)}, batch2 recheck 행: {len(r2)}, "
          f"둘 다 존재(같은 broad_tag로 재검증 트리거됨): {len(joined)}")
    only_b1 = keys_r1 - keys_r2
    only_b2 = keys_r2 - keys_r1
    print(f"batch1에서만 재검증 트리거됨: {len(only_b1)}건")
    print(f"batch2에서만 재검증 트리거됨: {len(only_b2)}건")
    print("(둘 중 한쪽에서만 트리거됐다는 건, 그 배치에서 primary_algorithm이나 GREEDY/DP/BITMASK/MATH "
          "점수 자체가 60점 문턱을 넘었는지 여부가 갈렸다는 뜻)")
    print()

    if not joined:
        print("공통 recheck 행이 없어 decision/candidate 비교를 건너뜁니다.")
        return

    decisions = ["KEEP_ORIGINAL", "REVIEW_NEEDED", "LIKELY_SPECIFIC_ALGORITHM"]
    decision_crosstab = Counter()
    candidate_match = 0
    broad_tag_counts = Counter()

    for key in joined:
        d1, d2 = r1[key]["recheck_decision"], r2[key]["recheck_decision"]
        decision_crosstab[(d1, d2)] += 1
        if r1[key]["recheck_best_candidate"] == r2[key]["recheck_best_candidate"]:
            candidate_match += 1
        broad_tag_counts[key[1]] += 1

    print(f"recheck_best_candidate 일치율: {candidate_match}/{len(joined)} "
          f"({candidate_match / len(joined) * 100:.1f}%)")
    print(f"broad_tag별 공통 recheck 건수: {dict(broad_tag_counts)}")
    print()
    print("recheck_decision 크로스탭 (batch1 x batch2):")
    header_line = "batch1\\batch2".ljust(28) + "".join(d.ljust(28) for d in decisions)
    print(header_line)
    for d1 in decisions:
        row_str = d1.ljust(28) + "".join(str(decision_crosstab[(d1, d2)]).ljust(28) for d2 in decisions)
        print(row_str)

    escalated = sum(1 for k in joined if r1[k]["recheck_decision"] == "KEEP_ORIGINAL"
                     and r2[k]["recheck_decision"] in ("REVIEW_NEEDED", "LIKELY_SPECIFIC_ALGORITHM"))
    deescalated = sum(1 for k in joined if r1[k]["recheck_decision"] in ("REVIEW_NEEDED", "LIKELY_SPECIFIC_ALGORITHM")
                        and r2[k]["recheck_decision"] == "KEEP_ORIGINAL")
    print()
    print(f"batch1 KEEP_ORIGINAL -> batch2 재검토 필요로 격상: {escalated}건")
    print(f"batch1 재검토 필요 -> batch2 KEEP_ORIGINAL로 완화: {deescalated}건")


if __name__ == "__main__":
    main()
